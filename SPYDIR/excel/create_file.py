import pandas as pd
import datetime
from SPYDIR.utils.helpers import get_img
from SPYDIR.financial.helpers import *

from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows

left_aligned_text = Alignment(horizontal="left")
# col
highlight = NamedStyle(name="col")
highlight.fill = PatternFill(fill_type="solid", fgColor="00FF00")
# Title
base_font = Font(name="Calibri", size=30)
title_style = NamedStyle(name="stock_title")
title_style.font = Font(
    name=base_font.name,
    size=base_font.size,
    bold=True,
    color="000000",
)


title_dict = {
    "font": Font(bold=True, color="FFFFFF"),
    "fill": PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),
    "alignment": Alignment(horizontal="left", vertical="center"),
}

body_dict = {
    "font": Font(bold=False, color="000000"),
    "fill": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    "alignment": Alignment(horizontal="left", vertical="center"),
}


def _cover_page(wb, stock_obj):
    ws = wb["Cover"]

    ws["B4"] = stock_obj.get("name", "")
    ws["B4"].style = "stock_title"
    ws["B7"] = stock_obj.get("ticker", "")
    ws["B7"].alignment = left_aligned_text
    ws["B8"] = datetime.date.today()
    ws["B8"].alignment = left_aligned_text
    ws.column_dimensions["B"].width = 50

    img_buffer = get_img(image_path=stock_obj.get("wiki", {}).get("img", ""))
    img = Image(img_buffer)
    ws.add_image(img, "F4")

    return wb


def _financials_page(wb, stock_obj):

    ws = wb["Financials"]
    related_dict = stock_obj.get("related", {})
    df = pd.DataFrame.from_dict(related_dict)
    df_to_excel(df, ws, header=True, index=False, startrow=20, startcol="B")

    return wb


def _statement_pages(wb, stock_obj):
    tabs = {
        "Income Statement": "income_statement",
        "Balance Sheet": "balance_sheet",
        "Cash Flow": "cash_flow",
        "Financials": "financial_ratios",
    }

    for key, value in tabs.items():

        ws = wb[key]
        state_dict = stock_obj.get("statements", {}).get(value, {})
        df = pd.DataFrame.from_dict(state_dict)
        df_to_excel(
            df,
            ws,
            header=True,
            index=True,
            startrow=2,
            startcol="B",
        )

    return wb


def df_to_excel(
    df,
    ws,
    index=True,
    header=True,
    body_dict=body_dict,
    title_dict=title_dict,
    startrow=1,
    startcol="A",
):
    col_offset = column_index_from_string(startcol) - 1

    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=index, header=header), start=startrow
    ):
        for c_idx, value in enumerate(row, start=1 + col_offset):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            is_header = header and r_idx == startrow
            is_index_col = index and c_idx == 1 + col_offset

            if is_header or is_index_col:
                cell.font = title_dict["font"]
                cell.fill = title_dict["fill"]
                cell.alignment = title_dict["alignment"]
            else:
                cell.font = body_dict["font"]
                cell.fill = body_dict["fill"]
                cell.alignment = body_dict["alignment"]
                if isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0_);("$"#,##0)'  # Accounting format

    default_width = 12
    index_width = default_width * 2

    for col_cells in ws.iter_cols(
        min_row=startrow,
        max_row=startrow,
        min_col=1 + col_offset,
        max_col=df.shape[1] + col_offset + (1 if index else 0),
    ):
        col_letter = col_cells[0].column_letter
        if index and col_cells[0].column == 1 + col_offset:
            ws.column_dimensions[col_letter].width = index_width
        else:
            ws.column_dimensions[col_letter].width = default_width
