from SPYDIR.stock_client import stock_client
from SPYDIR.excel.create_file import (
    _statement_pages,
    _financials_page,
    _cover_page,
    title_style,
    highlight,
)
from io import BytesIO
import copy
from openpyxl import load_workbook
import os


# Python
def get_excel(ticker, output=None):
    context = get_report_context(ticker)

    done = create_spreadsheet(context, output=output)
    if done:
        return True
    else:
        return False


def get_report_context(ticker):
    stock_obj = stock_client(ticker, arg="report")
    context = copy.deepcopy(stock_obj)

    ref_links = {
        "Website": context.get("website", ""),
        "IR Website": context.get("ir_website", ""),
        "Recent Filing": context.get("latest_report_url", ""),
        "Wikipedia": context.get("wiki", {"url": ""})["url"],
    }

    context["related_ticks"] = context.get("related", {})

    context["ref_links"] = ref_links
    context["ticker"] = ticker

    context["news"] = context.get("news", ["", "", "", "", ""])[0:4]

    return context


def create_spreadsheet(stock_obj, output=None):
    template_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "Stock_Template.xlsm"
    )  # xlsx, xltm
    wb = load_workbook(template_path, keep_vba=True)

    # Style
    wb.add_named_style(title_style)
    wb.add_named_style(highlight)

    wb = _cover_page(wb, stock_obj)
    wb = _statement_pages(wb, stock_obj)
    wb = _financials_page(wb, stock_obj)
    wb.active = wb.sheetnames.index("Cover")

    if isinstance(output, str):
        wb.save(output)
        print(f"Saved to: {output}")
        return True
    elif isinstance(output, BytesIO):
        wb.save(output)
        output.seek(0)
        return output
    else:
        print("Output Not Valid")
        return False
